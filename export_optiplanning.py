#!/usr/bin/env python3
"""
Export Optiplanning - Lanceur executable.
Genere les fichiers d'export depuis Outil_Material_Import.xlsm :
  1. TXT Optiplanning (8 colonnes tab-delimited)
  2. XML Plaques Nesting (boards pour SWOOD Nesting)
  3. XML Materiaux (materials complet pour SWOOD)
  4. XML Chants (edgebands pour SWOOD)
"""

import os
import sys
import uuid
import math
import re
from datetime import datetime
from dataclasses import dataclass, field
from typing import Optional, List
import xml.etree.ElementTree as ET
from xml.dom import minidom
from xml.sax.saxutils import escape as _xml_escape

# On embarque tout le code directement (pas d'import externe sauf openpyxl)
try:
    import openpyxl
except ImportError as exc:
    raise ImportError(
        "Module openpyxl manquant. Installer : pip install openpyxl"
    ) from exc


def _load_gui_dependencies():
    """Charge Tkinter/Pillow uniquement pour le lanceur graphique autonome.

    DestriCenter importe seulement les fonctions d export dans sa SPA
    WebView2. Ce chargement paresseux garde ces fonctions utilisables dans un
    bundle qui exclut volontairement Tkinter.
    """
    global tk, filedialog, ttk, Image, ImageTk
    import tkinter as tk
    from tkinter import filedialog, ttk
    from PIL import Image, ImageTk


def _xesc(val: str) -> str:
    """Echappe les caracteres speciaux XML pour les valeurs d'attributs."""
    return _xml_escape(val, {'"': '&quot;', "'": '&apos;'})


def _write_unique_export(output_dir: str, prefix: str, extension: str,
                         content: str) -> tuple:
    """Reserve puis ecrit un export sans collision inter-processus."""
    os.makedirs(output_dir, exist_ok=True)
    for _attempt in range(20):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        filename = f"{prefix}_{timestamp}_{uuid.uuid4().hex[:8]}{extension}"
        output_path = os.path.join(output_dir, filename)
        try:
            fd = os.open(
                output_path, os.O_WRONLY | os.O_CREAT | os.O_EXCL, 0o600)
        except FileExistsError:
            continue
        try:
            with os.fdopen(fd, "w", encoding="utf-8", newline="") as handle:
                handle.write(content)
                handle.flush()
                os.fsync(handle.fileno())
        except Exception:
            try:
                os.remove(output_path)
            except OSError:
                pass
            raise
        return filename, output_path
    raise FileExistsError("Impossible de reserver un nom d'export unique")


# ---------------------------------------------------------------------------
# Constantes XML SWOOD
# ---------------------------------------------------------------------------

APP_VERSION = "1.2"

SWOOD_XMLNS = "http://www.eficad.com//SWOODMat"
SWOOD_XSD = "http://www.w3.org/2001/XMLSchema"
SWOOD_XSI = "http://www.w3.org/2001/XMLSchema-instance"
SWOOD_VERSION = "2"


# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------

@dataclass
class MaterialSWOOD:
    """Un materiau SWOOD lu depuis le XLSM (page Materials)."""
    name: str = ""
    description: str = ""
    path: str = ""
    thickness: str = ""
    fiber_material: str = ""
    cost: str = ""
    density: str = ""
    color: str = ""
    transparency: str = ""
    texture: str = ""
    texture_direction: str = ""
    saw_stock: str = ""
    saw_reference: str = ""
    saw_fiber: str = ""
    fiber_speed_factor: str = ""
    fiber_angle_correction: str = ""
    material_type: str = ""
    material_costing_type: str = ""
    top_color: str = ""
    top_texture: str = ""
    top_texture_angle: str = ""
    top_texture_image_direction: str = ""
    bottom_color: str = ""
    bottom_texture: str = ""
    bottom_texture_angle: str = ""
    bottom_texture_image_direction: str = ""
    end_texture: str = ""
    sw_material: str = ""
    image: str = ""
    edge_band_list: str = ""
    laminate_impact: str = ""
    allow_thickness_calibration: str = ""
    min_thickness_calibration: str = ""
    machining_cost_factor: str = ""
    sw_texture_height: str = ""
    top_texture_height: str = ""
    bottom_texture_height: str = ""
    material_name_top: str = ""
    grain_direction_top: str = ""
    stock_offset_top: str = ""
    material_name_bottom: str = ""
    grain_direction_bottom: str = ""
    stock_offset_bottom: str = ""
    board_l: str = ""
    board_w: str = ""
    ref_fournisseur: str = ""
    fournisseur: str = ""
    finish: str = ""
    glass: str = ""
    # Champs calcules
    parametres: str = ""


@dataclass
class EdgeBandSWOOD:
    """Un chant SWOOD lu depuis le XLSM (page EdgeBands)."""
    name: str = ""
    id_val: str = ""
    description: str = ""
    path: str = ""
    cost: str = ""
    reference: str = ""
    thickness: str = ""
    color: str = ""
    image_path: str = ""
    creation_corps: str = ""
    stock_offset: str = ""
    width_min: str = ""
    width_max: str = ""
    width: str = ""
    force_stock_exclusion: str = ""
    shape_id: str = ""
    end_shape_id: str = ""
    use_mitre_cut: str = ""
    texture_height: str = ""
    eb_additional_shape_id: str = ""
    ebw_finish: str = ""
    finish: str = ""
    eb_supplier: str = ""


# ---------------------------------------------------------------------------
# Fonctions de calcul
# ---------------------------------------------------------------------------

def compute_saw_reference(name: str, thickness) -> str:
    """Calcule le SawReference par defaut : identique au Name.
    (Anciennement ajoutait ' XX mm' pour les melamine, ce qui creait
    une incoherence avec le Name dans le nesting.)
    """
    if not name:
        return ""
    return name


def _parse_finite_float(value):
    """Convertit value en float fini, ou None si absent/illisible/infini.

    Centralise la regle : une valeur source vide, non numerique ou non
    finie (NaN/Infini) ne doit jamais etre traitee comme un nombre exploitable.
    """
    if value is None:
        return None
    s = str(value).strip()
    if s == "":
        return None
    try:
        num = float(s)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(num):
        return None
    return num


def compute_parametres(board_l) -> str:
    # Sans longueur de plaque finie et positive, aucun gabarit Destribois
    # ne peut etre deduit : retourner '' plutot qu'une valeur inventee.
    num = _parse_finite_float(board_l)
    if num is None or num <= 0:
        return ""
    return "Destribois 5m" if num > 3200 else "Destribois"


def format_cost(cost) -> str:
    # Un cout absent, non numerique, non fini ou negatif est un cout
    # inconnu : retourner '' plutot que d'inventer un tarif par defaut.
    num = _parse_finite_float(cost)
    if num is None or num < 0:
        return ""
    return f"{num:.2f}"


def format_positive_number(value) -> str:
    """Convertit une cote en millimetres vers des metres (sans zeros
    inutiles), pour les attributs XML SWOOD (Length/Width/Thickness).

    Une cote absente, non numerique, non finie ou non strictement positive
    n'a pas de sens physique : retourner '' plutot qu'une dimension inventee.
    """
    num = _parse_finite_float(value)
    if num is None or num <= 0:
        return ""
    meters = num / 1000.0
    return f"{meters:g}"


def _safe_str(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if s.endswith(".0"):
        try:
            int_val = int(float(s))
            if float(s) == int_val:
                return str(int_val)
        except (ValueError, OverflowError):
            pass
    return s


def _bool_str(value) -> str:
    """Convertit une valeur en 'true'/'false' pour XML SWOOD."""
    if value is None:
        return "false"
    s = str(value).strip().lower()
    if s in ("1", "true", "yes", "oui"):
        return "true"
    return "false"


def _grain_direction(name: str) -> str:
    """Determine la direction du grain a partir du nom du materiau."""
    n = name.lower()
    if any(x in n for x in ("h1", "h3", "f1", "f2", "f3", "f4")):
        # Les decors bois/textures ont generalement un sens horizontal
        if "melamine-h" in n or "melamine-f" in n:
            return "Horizontal"
    return "None"


# ---------------------------------------------------------------------------
# Lecture XLSM - Page Materials (complete, 49 colonnes)
# ---------------------------------------------------------------------------

def _resolve_cell(ws, row, col):
    """Lit une cellule et resout les formules simples (=XX123)."""
    import re as _re
    val = ws.cell(row=row, column=col).value
    if isinstance(val, str) and val.startswith("="):
        m = _re.match(r"^=([A-Z]{1,3})(\d+)$", val)
        if m:
            ref_col_str = m.group(1)
            ref_row = int(m.group(2))
            ref_col = 0
            for ch in ref_col_str:
                ref_col = ref_col * 26 + (ord(ch) - ord('A') + 1)
            ref_val = ws.cell(row=ref_row, column=ref_col).value
            if isinstance(ref_val, str) and ref_val.startswith("="):
                return val
            return ref_val
    return val


def read_all_materials_from_xlsm(xlsm_path: str, log_func=print) -> List[MaterialSWOOD]:
    """Lit TOUTES les colonnes de la page Materials (49 colonnes)."""
    log_func(f"Lecture de : {os.path.basename(xlsm_path)} (Materials - complet)")
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True, data_only=False)
    ws = wb["Materials"]
    materials = []
    for row in range(5, ws.max_row + 1):
        name = _resolve_cell(ws, row, 1)
        if not name or str(name).strip() == "":
            continue
        mat = MaterialSWOOD(
            name=str(name).strip(),
            description=_safe_str(_resolve_cell(ws, row, 2)),
            path=_safe_str(_resolve_cell(ws, row, 3)),
            thickness=_safe_str(_resolve_cell(ws, row, 4)),
            fiber_material=_safe_str(_resolve_cell(ws, row, 5)),
            cost=_safe_str(_resolve_cell(ws, row, 6)),
            density=_safe_str(_resolve_cell(ws, row, 7)),
            color=_safe_str(_resolve_cell(ws, row, 8)),
            transparency=_safe_str(_resolve_cell(ws, row, 9)),
            texture=_safe_str(_resolve_cell(ws, row, 10)),
            texture_direction=_safe_str(_resolve_cell(ws, row, 11)),
            saw_stock=_safe_str(_resolve_cell(ws, row, 12)),
            saw_reference=_safe_str(_resolve_cell(ws, row, 13)),
            saw_fiber=_safe_str(_resolve_cell(ws, row, 14)),
            fiber_speed_factor=_safe_str(_resolve_cell(ws, row, 15)),
            fiber_angle_correction=_safe_str(_resolve_cell(ws, row, 16)),
            material_type=_safe_str(_resolve_cell(ws, row, 17)),
            material_costing_type=_safe_str(_resolve_cell(ws, row, 18)),
            top_color=_safe_str(_resolve_cell(ws, row, 19)),
            top_texture=_safe_str(_resolve_cell(ws, row, 20)),
            top_texture_angle=_safe_str(_resolve_cell(ws, row, 21)),
            top_texture_image_direction=_safe_str(_resolve_cell(ws, row, 22)),
            bottom_color=_safe_str(_resolve_cell(ws, row, 23)),
            bottom_texture=_safe_str(_resolve_cell(ws, row, 24)),
            bottom_texture_angle=_safe_str(_resolve_cell(ws, row, 25)),
            bottom_texture_image_direction=_safe_str(_resolve_cell(ws, row, 26)),
            end_texture=_safe_str(_resolve_cell(ws, row, 27)),
            sw_material=_safe_str(_resolve_cell(ws, row, 28)),
            image=_safe_str(_resolve_cell(ws, row, 29)),
            edge_band_list=_safe_str(_resolve_cell(ws, row, 30)),
            laminate_impact=_safe_str(_resolve_cell(ws, row, 31)),
            allow_thickness_calibration=_safe_str(_resolve_cell(ws, row, 32)),
            min_thickness_calibration=_safe_str(_resolve_cell(ws, row, 33)),
            machining_cost_factor=_safe_str(_resolve_cell(ws, row, 34)),
            sw_texture_height=_safe_str(_resolve_cell(ws, row, 35)),
            top_texture_height=_safe_str(_resolve_cell(ws, row, 36)),
            bottom_texture_height=_safe_str(_resolve_cell(ws, row, 37)),
            material_name_top=_safe_str(_resolve_cell(ws, row, 38)),
            grain_direction_top=_safe_str(_resolve_cell(ws, row, 39)),
            stock_offset_top=_safe_str(_resolve_cell(ws, row, 40)),
            material_name_bottom=_safe_str(_resolve_cell(ws, row, 41)),
            grain_direction_bottom=_safe_str(_resolve_cell(ws, row, 42)),
            stock_offset_bottom=_safe_str(_resolve_cell(ws, row, 43)),
            board_l=_safe_str(_resolve_cell(ws, row, 44)),
            board_w=_safe_str(_resolve_cell(ws, row, 45)),
            ref_fournisseur=_safe_str(_resolve_cell(ws, row, 46)),
            fournisseur=_safe_str(_resolve_cell(ws, row, 47)),
            finish=_safe_str(_resolve_cell(ws, row, 48)),
            glass=_safe_str(_resolve_cell(ws, row, 49)),
        )
        mat.parametres = compute_parametres(mat.board_l)
        if not mat.saw_reference:
            mat.saw_reference = compute_saw_reference(mat.name, mat.thickness)
        materials.append(mat)
    wb.close()
    log_func(f"{len(materials)} materiaux lus (49 colonnes)")
    return materials


def read_materials_from_xlsm(xlsm_path: str, log_func=print) -> list:
    """Lit les colonnes essentielles de la page Materials (export TXT)."""
    log_func(f"Lecture de : {os.path.basename(xlsm_path)}")
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True, data_only=True)
    ws = wb["Materials"]
    materials = []
    for row in range(5, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if not name or str(name).strip() == "":
            continue
        mat = MaterialSWOOD(
            name=str(name).strip(),
            thickness=_safe_str(ws.cell(row=row, column=4).value),
            fiber_material=_safe_str(ws.cell(row=row, column=5).value),
            cost=ws.cell(row=row, column=6).value,
            board_l=_safe_str(ws.cell(row=row, column=44).value),
            board_w=_safe_str(ws.cell(row=row, column=45).value),
            ref_fournisseur=_safe_str(ws.cell(row=row, column=46).value),
        )
        mat.saw_reference = compute_saw_reference(mat.name, mat.thickness)
        mat.parametres = compute_parametres(mat.board_l)
        mat.cost = format_cost(mat.cost)
        materials.append(mat)
    wb.close()
    log_func(f"{len(materials)} materiaux lus")
    return materials


# ---------------------------------------------------------------------------
# Lecture XLSM - Page EdgeBands
# ---------------------------------------------------------------------------

def read_edgebands_from_xlsm(xlsm_path: str, log_func=print) -> List[EdgeBandSWOOD]:
    """Lit la page EdgeBands du XLSM (23 colonnes)."""
    log_func(f"Lecture de : {os.path.basename(xlsm_path)} (EdgeBands)")
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True, data_only=True)

    if "EdgeBands" not in wb.sheetnames:
        log_func("ERREUR : Page 'EdgeBands' introuvable dans le XLSM.")
        wb.close()
        return []

    ws = wb["EdgeBands"]
    edgebands = []
    for row in range(5, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if not name or str(name).strip() == "":
            continue
        eb = EdgeBandSWOOD(
            name=str(name).strip(),
            id_val=_safe_str(ws.cell(row=row, column=2).value),
            description=_safe_str(ws.cell(row=row, column=3).value),
            path=_safe_str(ws.cell(row=row, column=4).value),
            cost=_safe_str(ws.cell(row=row, column=5).value),
            reference=_safe_str(ws.cell(row=row, column=6).value),
            thickness=_safe_str(ws.cell(row=row, column=7).value),
            color=_safe_str(ws.cell(row=row, column=8).value),
            image_path=_safe_str(ws.cell(row=row, column=9).value),
            creation_corps=_safe_str(ws.cell(row=row, column=10).value),
            stock_offset=_safe_str(ws.cell(row=row, column=11).value),
            width_min=_safe_str(ws.cell(row=row, column=12).value),
            width_max=_safe_str(ws.cell(row=row, column=13).value),
            width=_safe_str(ws.cell(row=row, column=14).value),
            force_stock_exclusion=_safe_str(ws.cell(row=row, column=15).value),
            shape_id=_safe_str(ws.cell(row=row, column=16).value),
            end_shape_id=_safe_str(ws.cell(row=row, column=17).value),
            use_mitre_cut=_safe_str(ws.cell(row=row, column=18).value),
            texture_height=_safe_str(ws.cell(row=row, column=19).value),
            eb_additional_shape_id=_safe_str(ws.cell(row=row, column=20).value),
            ebw_finish=_safe_str(ws.cell(row=row, column=21).value),
            finish=_safe_str(ws.cell(row=row, column=22).value),
            eb_supplier=_safe_str(ws.cell(row=row, column=23).value),
        )
        edgebands.append(eb)
    wb.close()
    log_func(f"{len(edgebands)} chants lus")
    return edgebands


# ---------------------------------------------------------------------------
# Utilitaire XML
# ---------------------------------------------------------------------------

def _pretty_xml(root_element: ET.Element) -> str:
    """Genere un XML proprement indente avec declaration UTF-8."""
    rough = ET.tostring(root_element, encoding="unicode")
    parsed = minidom.parseString(rough)
    pretty = parsed.toprettyxml(indent="  ", encoding=None)
    # Supprimer la ligne <?xml ?> generee par minidom (on la met nous-meme)
    lines = pretty.split("\n")
    # minidom ajoute <?xml version="1.0" ?> qu'on remplace par notre version
    if lines and lines[0].startswith("<?xml"):
        lines[0] = '<?xml version="1.0" encoding="utf-8"?>'
    # Nettoyer les lignes vides en trop
    cleaned = []
    for line in lines:
        if line.strip():
            cleaned.append(line)
    return "\n".join(cleaned)


def _create_swood_root() -> ET.Element:
    """Cree l'element racine <SWOODMat> avec les bons namespaces."""
    root = ET.Element("SWOODMat")
    root.set("xmlns:xsd", SWOOD_XSD)
    root.set("xmlns:xsi", SWOOD_XSI)
    root.set("Version", SWOOD_VERSION)
    root.set("xmlns", SWOOD_XMLNS)
    return root


def _validate_xml_document(content: str) -> None:
    """Refuse d'ecrire un export incomplet ou XML mal forme."""
    try:
        ET.fromstring(content)
    except ET.ParseError as exc:
        raise ValueError("Export XML SWOOD mal forme : %s" % exc) from exc


# ---------------------------------------------------------------------------
# EXPORT 1 : TXT Optiplanning (existant)
# ---------------------------------------------------------------------------

def generate_optiplanning_lines(materials: list) -> list:
    lines = []
    for mat in materials:
        cols = [
            mat.saw_reference,
            mat.board_l,
            mat.board_w,
            mat.thickness,
            mat.fiber_material,
            mat.cost,
            mat.parametres,
            mat.ref_fournisseur,
        ]
        lines.append("\t".join(cols))
    return lines


def export_optiplanning_txt(xlsm_path: str, output_dir: str = None, log_func=print) -> str:
    """Export TXT Optiplanning (8 colonnes tab-delimited)."""
    materials = read_materials_from_xlsm(xlsm_path, log_func)
    if not materials:
        log_func("ERREUR : Aucun materiau lu.")
        return ""

    lines = generate_optiplanning_lines(materials)

    if output_dir is None:
        output_dir = os.path.dirname(os.path.abspath(xlsm_path))

    filename, output_path = _write_unique_export(
        output_dir, "Materiaux_a_importer_Optiplanning", ".txt",
        "\n".join(lines))

    count_5m = sum(1 for m in materials if m.parametres == "Destribois 5m")
    count_missing_cost = sum(1 for m in materials if not m.cost)
    count_no_ref = sum(1 for m in materials if not m.ref_fournisseur)

    log_func(f"Fichier cree : {filename}")
    log_func(f"  {len(lines)} lignes")
    log_func(f"  {count_5m} lignes 'Destribois 5m'")
    if count_missing_cost:
        log_func(f"  {count_missing_cost} lignes sans cout (valeur absente)")
    if count_no_ref:
        log_func(f"  {count_no_ref} lignes sans ref fournisseur")

    return output_path


# ---------------------------------------------------------------------------
# EXPORT 2 : XML Plaques Nesting (structure identique a Structure.xml)
# ---------------------------------------------------------------------------

def export_xml_boards_nesting(xlsm_path: str, output_dir: str = None, log_func=print) -> str:
    """Export XML plaques pour SWOOD Nesting.

    Genere le XML en texte brut (meme format que la macro VBA) pour
    une compatibilite maximale avec l'import SWOOD.
    Structure : <SWOODMat> -> <Boards> -> <Board ... />
    Dimensions en mm (identique au fichier de reference Structure_plaques_nesting.xml).
    """
    materials = read_all_materials_from_xlsm(xlsm_path, log_func)
    if not materials:
        log_func("ERREUR : Aucun materiau lu.")
        return ""

    log_func(f"Generation XML Plaques Nesting...")

    # Lire l'entete XML depuis le XLSM (identique a la macro VBA)
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True, data_only=False)
    ws = wb["Materials"]
    xml_line1 = str(ws.cell(row=1, column=1).value or '<?xml version="1.0" encoding="utf-8"?>')
    xml_line2 = str(ws.cell(row=2, column=1).value or "")
    wb.close()

    # Construction du XML en texte brut (meme format que la macro VBA)
    txt = xml_line1 + "\r\n" + xml_line2
    txt += "\r\n\t<Boards>"

    count = 0
    ignored = 0
    grain_count = 0
    idx = 0
    for mat in materials:
        # Cotes obligatoires (mm -> m). Une plaque sans longueur, largeur
        # ou epaisseur finie et positive n'est pas nestable : elle est
        # ignoree plutot que completee par une dimension inventee.
        length_val = format_positive_number(mat.board_l)
        width_val = format_positive_number(mat.board_w)
        thick_val = format_positive_number(mat.thickness)
        if not (length_val and width_val and thick_val):
            ignored += 1
            continue

        idx += 1
        count += 1

        # GrainDirection
        grain = "Horizontal" if mat.fiber_material == "1" else "None"
        if grain == "Horizontal":
            grain_count += 1

        # Materials = SawReference ou Name
        materials_val = mat.saw_reference if mat.saw_reference else mat.name

        # Construction du <Board ... /> en texte brut (self-closing)
        txt += "\r\n\t\t<Board"
        txt += f' Name="{_xesc(mat.name)}"'
        txt += f' Description="{_xesc(mat.description)}"'
        txt += f' Path="{_xesc(mat.path)}"'
        txt += f' BoardType="Panel"'
        txt += f' Length="{length_val}"'
        txt += f' Width="{width_val}"'
        txt += f' Thickness="{thick_val}"'
        txt += f' GrainDirection="{grain}"'
        # Cost = surface m2 x prix/m2, omis (pas invente) si le prix
        # source est absent, non numerique, non fini ou negatif.
        cost_m2_str = format_cost(mat.cost)
        if cost_m2_str:
            cost_plaque = float(length_val) * float(width_val) * float(cost_m2_str)
            txt += f' Cost="{cost_plaque:.2f}"'
        txt += f' MaterialID="0"'
        txt += f' Reference="{_xesc(mat.ref_fournisseur)}"'
        txt += f' Supplier="{_xesc(mat.fournisseur)}"'
        txt += f' SupplierReference="{_xesc(mat.ref_fournisseur)}"'
        txt += f' NestingCorner="Lower_Left"'
        txt += f' NestingDirection="X"'
        txt += f' NestingUniformCollar="0"'
        txt += f' DefaultNestPriority="1"'
        txt += f' TopMaterial=""'
        txt += f' TopGrainAngle="NaN"'
        txt += f' BottomMaterial=""'
        txt += f' BottomGrainAngle="NaN"'
        txt += f' CanFlipTopBottom="false"'
        txt += f' LibraryUUID="{uuid.uuid4()}"'
        txt += f' ID="{idx}"'
        txt += f' ForBoardEstimation="true"'
        txt += f' Materials="{_xesc(materials_val)}"'
        txt += " />"

    txt += "\r\n\t</Boards>"
    txt += "\r\n</SWOODMat>"

    # Ecriture du fichier
    if output_dir is None:
        output_dir = os.path.dirname(os.path.abspath(xlsm_path))

    filename, output_path = _write_unique_export(
        output_dir, "Plaques_Nesting", ".xml", txt)

    log_func(f"Fichier cree : {filename}")
    log_func(f"  {count} plaques exportees")
    log_func(f"  {grain_count} plaques avec grain horizontal")
    log_func(f"  {ignored} plaques ignorees (cotes absentes ou non finies)")

    return output_path


# ---------------------------------------------------------------------------
# EXPORT 3 & 4 : Reproduction fidele de la macro VBA du XLSM
# La macro lit les tags en row 3 et headers en row 4 pour construire le XML.
# Chaque colonne est traitee selon son tag :
#   ""           -> attribut simple du noeud <Material ...> ou <EdgeBand ...>
#   "Properties" -> ouvre un bloc <Properties>, ecrit <Property Name="H" Value="V" />
#   "Property"   -> ecrit un <Property Name="H" Value="V" /> dans le bloc courant
#   "/Properties"-> ecrit la property puis ferme </Properties>
#   "Layers"     -> ouvre un bloc <Layers>, commence un <Layer H="V" ...>
#   "Layer"      -> continue les attributs du <Layer> ou commence un nouveau
#   "/Layer"     -> ferme le <Layer ... /> courant
#   "/Layers"    -> ferme le <Layer> puis </Layers>
# ---------------------------------------------------------------------------

_EDGEBAND_VARIANT_PATHS = ("1x23", "2x23", "1x43", "2x43")
_EDGEBAND_WIDTH_HEADERS = ("WidthMin", "WidthMax", "Width")


def _edgeband_variant_base(name: str, path: str) -> Optional[str]:
    prefix = path + " - "
    value = _format_cell_value(name)
    if not value.startswith(prefix):
        return None
    base = value[len(prefix):].strip()
    return base or None


def _edgeband_archive_base(name: str, thickness) -> Optional[tuple]:
    """Retourne le decor et son epaisseur depuis un nom d'archive.

    Les quatre listes actives sont des projections des lignes ``Archives``.
    Le suffixe du nom est controle avec la colonne Thickness afin de ne pas
    fabriquer une variante depuis une ligne ambigue ou mal etiquetee.
    """
    value = _format_cell_value(name)
    match = re.match(r"^(.+?)\s+-\s+([12])(?:\.0+)?\s*mm\s*$", value,
                     flags=re.IGNORECASE)
    if not match:
        return None
    declared = _format_cell_value(thickness)
    if declared.endswith(".0"):
        declared = declared[:-2]
    archive_thickness = match.group(2)
    if declared != archive_thickness:
        return None
    return match.group(1).strip(), archive_thickness


def _normalise_edgeband_rows(rows: list, headers: list) -> list:
    """Reconstruit les variantes actives et reserve les ID des Archives.

    Le classeur contient une ligne canonique par decor/epaisseur sous
    ``Path=Archives`` et quatre listes actives derivees (1x23, 2x23, 1x43,
    2x43). Repartir de ces lignes canoniques evite les decalages de blocs et
    recree une variante manquante. Les lignes Archives restent byte-for-byte
    equivalentes du point de vue de leurs valeurs ; seules les lignes actives
    derivees recoivent de nouveaux ID.
    """
    indexes = {header: index for index, header in enumerate(headers) if header}
    required = {"Name", "ID", "Path", "Reference", "Thickness"}
    if not required <= set(indexes):
        return rows

    name_index = indexes["Name"]
    id_index = indexes["ID"]
    path_index = indexes["Path"]
    reference_index = indexes["Reference"]
    thickness_index = indexes["Thickness"]
    width_indexes = [indexes[name] for name in _EDGEBAND_WIDTH_HEADERS
                     if name in indexes]

    archives = []
    active_rows = []
    archive_sources = {"1": {}, "2": {}}
    archive_order = {"1": [], "2": []}
    for row in rows:
        path = _format_cell_value(row[path_index])
        if path != "Archives":
            active_rows.append(row)
            continue
        archives.append(row)
        parsed = _edgeband_archive_base(
            row[name_index], row[thickness_index])
        if not parsed:
            continue
        base, thickness = parsed
        if base not in archive_sources[thickness]:
            archive_sources[thickness][base] = row
            archive_order[thickness].append(base)

    existing_variants = {}
    active_order_by_path = {path: [] for path in _EDGEBAND_VARIANT_PATHS}
    managed_rows = set()
    for row in active_rows:
        path = _format_cell_value(row[path_index])
        if path not in _EDGEBAND_VARIANT_PATHS:
            continue
        base = _edgeband_variant_base(row[name_index], path)
        thickness = path.split("x", 1)[0]
        if not base or base not in archive_sources[thickness]:
            continue
        key = (path, base)
        managed_rows.add(id(row))
        if key not in existing_variants:
            existing_variants[key] = row
            active_order_by_path[path].append(base)

    order_by_thickness = {}
    for thickness in ("1", "2"):
        order = []
        seen = set()
        same_thickness_paths = [
            path for path in _EDGEBAND_VARIANT_PATHS
            if path.startswith(thickness + "x")]
        fallback_paths = [
            path for path in _EDGEBAND_VARIANT_PATHS
            if path not in same_thickness_paths]
        for path in same_thickness_paths + fallback_paths:
            for base in active_order_by_path[path]:
                if (base in archive_sources[thickness]
                        and base not in seen):
                    seen.add(base)
                    order.append(base)
        for base in archive_order[thickness]:
            if base not in seen:
                seen.add(base)
                order.append(base)
        order_by_thickness[thickness] = order

    variants_by_path = {}
    for path in _EDGEBAND_VARIANT_PATHS:
        thickness = path.split("x", 1)[0]
        variants = []
        for base in order_by_thickness[thickness]:
            source = list(archive_sources[thickness][base])
            current = existing_variants.get((path, base))
            if current is not None:
                for index in width_indexes:
                    source[index] = current[index]
            variant_name = path + " - " + base
            source[name_index] = variant_name
            source[id_index] = None
            source[path_index] = path
            source[reference_index] = variant_name
            variants.append(source)
        variants_by_path[path] = variants

    normalised = []
    emitted_paths = set()
    for row in active_rows:
        path = _format_cell_value(row[path_index])
        if id(row) in managed_rows:
            if path not in emitted_paths:
                normalised.extend((variant, "generated")
                                  for variant in variants_by_path[path])
                emitted_paths.add(path)
            continue
        normalised.append((list(row), "active"))
    for path in _EDGEBAND_VARIANT_PATHS:
        if path not in emitted_paths and variants_by_path[path]:
            normalised.extend((variant, "generated")
                              for variant in variants_by_path[path])
    normalised.extend((list(row), "archive") for row in archives)

    archive_ids = []
    for row, kind in normalised:
        if kind == "archive":
            identifier = _format_cell_value(row[id_index])
            if not identifier or identifier in archive_ids:
                raise ValueError(
                    "Les ID des chants Archives doivent etre renseignes et uniques")
            archive_ids.append(identifier)

    used_ids = set(archive_ids)
    needs_id = []
    active_names = set()
    for row, kind in normalised:
        if kind == "archive":
            continue
        name = _format_cell_value(row[name_index])
        if name in active_names:
            raise ValueError("Nom de chant actif duplique : " + name)
        active_names.add(name)
        identifier = _format_cell_value(row[id_index])
        if kind == "generated" or not identifier or identifier in used_ids:
            needs_id.append(row)
            row[id_index] = None
        else:
            used_ids.add(identifier)

    numeric_ids = [int(value) for value in used_ids if value.isdigit()]
    next_id = max(numeric_ids, default=0) + 1
    for row in needs_id:
        while str(next_id) in used_ids:
            next_id += 1
        row[id_index] = next_id
        used_ids.add(str(next_id))
        next_id += 1

    return [row for row, _kind in normalised]

def _format_cell_value(val) -> str:
    """Formate une valeur de cellule comme la macro VBA :
    - Remplace les virgules par des points
    - TRUE/FALSE en minuscules
    - Retourne une string."""
    if val is None:
        return ""
    s = str(val).strip()
    s = s.replace(",", ".")
    if s.upper() in ("TRUE", "FALSE"):
        s = s.lower()
    return s


def _export_vba_xml_sheet(xlsm_path: str, sheet_name: str, output_dir: str = None,
                          output_prefix: str = "Export", log_func=print) -> str:
    """Reproduit exactement la logique de la macro VBA SaveTextToFile pour une sheet.

    Lit les cellules A1, A2 (entete XML), row 3 (tags), row 4 (headers),
    puis parcourt les donnees (row 5+) en construisant le XML exactement comme
    le fait la macro VBA du XLSM.
    """
    log_func(f"Lecture de : {os.path.basename(xlsm_path)} ({sheet_name})")
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True, data_only=False)
    ws = wb[sheet_name]

    lastrow = ws.max_row
    lastcol = ws.max_column

    # Lire l'entete XML depuis A1 et A2
    xml_line1 = str(ws.cell(row=1, column=1).value or "<?xml version=\"1.0\" encoding=\"utf-8\"?>")
    xml_line2 = str(ws.cell(row=2, column=1).value or "")

    # Lire tags (row 3) et headers (row 4)
    tags = []
    headers = []
    for j in range(1, lastcol + 1):
        tag = ws.cell(row=3, column=j).value
        tags.append(str(tag).strip() if tag else "")
        header = ws.cell(row=4, column=j).value
        headers.append(str(header).strip() if header else "")

    # Alias : Materials -> Material, EdgeBands -> EdgeBand
    if sheet_name == "Materials":
        obj_alias = "Material"
    elif sheet_name == "EdgeBands":
        obj_alias = "EdgeBand"
    else:
        obj_alias = sheet_name.rstrip("s")

    # Construction du texte XML (reproduction fidele de la macro VBA)
    txt = xml_line1 + "\r\n" + xml_line2
    txt += "\r\n\t<" + sheet_name + ">"

    # Helper pour resoudre les formules simples de type =COL_LETTRE+LIGNE
    # (ex: =AT5 -> lire la valeur de la colonne AT ligne 5)
    import re as _re
    def _resolve_cell_value(ws, row, col):
        """Lit une cellule et resout les formules simples (=XX123)."""
        val = ws.cell(row=row, column=col).value
        if isinstance(val, str) and val.startswith("="):
            # Formule simple de type =AT5, =A5, =B12, etc.
            m = _re.match(r"^=([A-Z]{1,3})(\d+)$", val)
            if m:
                ref_col_str = m.group(1)
                ref_row = int(m.group(2))
                ref_col = 0
                for ch in ref_col_str:
                    ref_col = ref_col * 26 + (ord(ch) - ord('A') + 1)
                ref_val = ws.cell(row=ref_row, column=ref_col).value
                # Verifier que la ref n'est pas aussi une formule (eviter boucle)
                if isinstance(ref_val, str) and ref_val.startswith("="):
                    return val  # retourner la formule telle quelle
                return ref_val
        return val

    source_rows = []
    for i in range(5, lastrow + 1):
        # Verifier que la ligne a un nom (col 1)
        name_val = _resolve_cell_value(ws, i, 1)
        if not name_val or str(name_val).strip() == "":
            continue
        source_rows.append([
            _resolve_cell_value(ws, i, column)
            for column in range(1, lastcol + 1)
        ])

    if sheet_name == "EdgeBands":
        source_rows = _normalise_edgeband_rows(source_rows, headers)

    count = len(source_rows)
    path_index = headers.index("Path") if "Path" in headers else None
    for row_values in source_rows:
        active_edgeband = (
            sheet_name == "EdgeBands"
            and path_index is not None
            and _format_cell_value(row_values[path_index]) != "Archives"
        )

        # Debut du noeud objet
        obj_txt = "\r\n\t\t<" + obj_alias
        needs_close_tag = False  # True si on a ouvert un sous-noeud (Layers/Properties)
        in_properties = False  # True si on est dans un bloc <Properties>
        in_layers = False  # True si on est dans un bloc <Layers>

        column_order = list(range(lastcol))
        if active_edgeband:
            # Toutes les colonnes sans tag de sous-noeud sont des attributs
            # EdgeBand, meme lorsqu'elles se trouvent apres /Properties dans
            # le classeur. Les ecrire d'abord les garde dans la balise ouvrante.
            column_order = (
                [j for j in column_order if tags[j] == ""]
                + [j for j in column_order if tags[j] != ""]
            )

        for j in column_order:
            tag = tags[j]
            header = headers[j]
            raw_val = row_values[j]
            cur_val = _format_cell_value(raw_val)

            # SWOOD attend FiberAngleCorrection en radians, le XLSM stocke des degres
            if header == "FiberAngleCorrection" and cur_val != "":
                try:
                    cur_val = str(math.radians(float(cur_val)))
                except ValueError:
                    pass

            # Pour les balises de fermeture, on doit toujours les traiter
            # meme si la valeur est vide
            if tag == "/Properties":
                if cur_val != "":
                    # Si Properties n'a pas ete ouvert (BOARDL vide etc.),
                    # il faut l'ouvrir maintenant avant d'ecrire la Property
                    if not in_properties:
                        if not needs_close_tag:
                            obj_txt += ">"
                            needs_close_tag = True
                        obj_txt += "\r\n\t\t\t<Properties>"
                        in_properties = True
                    obj_txt += "\r\n\t\t\t\t<Property Name=\"" + header + "\" Value=\"" + _xesc(cur_val) + "\" />"
                if in_properties:
                    obj_txt += "\r\n\t\t\t</Properties>"
                    in_properties = False
                continue

            if tag == "/Layers":
                if in_layers:
                    if cur_val != "":
                        obj_txt += " " + header + "=\"" + _xesc(cur_val) + "\" />"
                    obj_txt += "\r\n\t\t\t</Layers>"
                    in_layers = False
                continue

            if tag == "/Layer":
                if in_layers:
                    if cur_val != "":
                        obj_txt += " " + header + "=\"" + _xesc(cur_val) + "\""
                    obj_txt += " />"
                continue

            if cur_val == "":
                continue

            if header == "":
                continue

            if tag == "":
                # Attribut simple
                obj_txt += " " + header + "=\"" + _xesc(cur_val) + "\""

            elif tag == "Properties":
                # Ouvrir le noeud objet (>) et commencer un bloc Properties
                if not needs_close_tag:
                    obj_txt += ">"
                    needs_close_tag = True
                obj_txt += "\r\n\t\t\t<Properties>"
                obj_txt += "\r\n\t\t\t\t<Property Name=\"" + header + "\" Value=\"" + _xesc(cur_val) + "\" />"
                in_properties = True

            elif tag == "Property":
                # Si Properties n'a pas ete ouvert (colonne Properties/BOARDL vide),
                # il faut l'ouvrir maintenant
                if not in_properties:
                    if not needs_close_tag:
                        obj_txt += ">"
                        needs_close_tag = True
                    obj_txt += "\r\n\t\t\t<Properties>"
                    in_properties = True
                obj_txt += "\r\n\t\t\t\t<Property Name=\"" + header + "\" Value=\"" + _xesc(cur_val) + "\" />"

            elif tag == "Layers":
                # Ouvrir le noeud objet (>) et commencer un bloc Layers
                if not needs_close_tag:
                    obj_txt += ">"
                    needs_close_tag = True
                obj_txt += "\r\n\t\t\t<Layers>"
                obj_txt += "\r\n\t\t\t\t<Layer " + header + "=\"" + _xesc(cur_val) + "\""
                in_layers = True

            elif tag == "Layer":
                # Verifier si le tag precedent etait /Layer -> nouveau Layer
                prev_tag = tags[j - 1] if j > 0 else ""
                if prev_tag == "/Layer":
                    obj_txt += "\r\n\t\t\t\t<Layer " + header + "=\"" + _xesc(cur_val) + "\""
                else:
                    obj_txt += " " + header + "=\"" + _xesc(cur_val) + "\""

        # Fermeture du noeud objet
        if needs_close_tag:
            # Le noeud a des sous-elements (Properties/Layers) -> fermeture explicite
            obj_txt += "\r\n\t\t</" + obj_alias + ">"
        else:
            # Le noeud n'a que des attributs -> self-closing />
            obj_txt += " />"

        txt += obj_txt

    txt += "\r\n\t</" + sheet_name + ">"
    wb.close()

    return txt, count


def export_xml_materials(xlsm_path: str, output_dir: str = None, log_func=print) -> str:
    """Export XML materiaux complet pour SWOOD.

    Reproduit exactement la macro VBA du XLSM en parcourant les 2 sheets
    (Materials + EdgeBands) et en utilisant les tags row 3 / headers row 4
    pour construire la structure XML identique.
    """
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True, data_only=False)
    ws = wb["Materials"]
    xml_line1 = str(ws.cell(row=1, column=1).value or "<?xml version=\"1.0\" encoding=\"utf-8\"?>")
    xml_line2 = str(ws.cell(row=2, column=1).value or "")
    wb.close()

    log_func(f"Generation XML SWOOD Materiaux (reproduction macro VBA)...")

    # Construire le XML pour Materials
    mat_txt, mat_count = _export_vba_xml_sheet(xlsm_path, "Materials", log_func=log_func)
    log_func(f"  {mat_count} materiaux lus")

    # Construire le XML pour EdgeBands
    eb_txt, eb_count = _export_vba_xml_sheet(xlsm_path, "EdgeBands", log_func=log_func)
    log_func(f"  {eb_count} chants lus")

    # Assembler le fichier final : entete + Materials + EdgeBands + fermeture
    # La macro VBA concatene les 2 sheets dans le meme fichier
    full_xml = xml_line1 + "\r\n" + xml_line2 + "\r\n"
    # Extraire le contenu apres l'entete (a partir de \r\n\t<Materials>)
    mat_body = mat_txt.split("\r\n", 2)[2] if "\r\n" in mat_txt else ""
    eb_body = eb_txt.split("\r\n", 2)[2] if "\r\n" in eb_txt else ""
    full_xml += mat_body + "\r\n" + eb_body
    full_xml += "\r\n</SWOODMat>"
    _validate_xml_document(full_xml)

    # Ecriture du fichier
    if output_dir is None:
        output_dir = os.path.dirname(os.path.abspath(xlsm_path))

    filename, output_path = _write_unique_export(
        output_dir, "Import_Swood_Materiaux", ".xml", full_xml)

    log_func(f"Fichier cree : {filename}")
    log_func(f"  Total : {mat_count} materiaux + {eb_count} chants")

    return output_path


# ---------------------------------------------------------------------------
# EXPORT 4 : XML Chants / EdgeBands seuls
# ---------------------------------------------------------------------------

def export_xml_edgebands(xlsm_path: str, output_dir: str = None, log_func=print) -> str:
    """Export XML chants seuls pour SWOOD.

    Reproduit la macro VBA du XLSM uniquement pour la sheet EdgeBands.
    """
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True, data_only=True)
    ws = wb["EdgeBands"]
    xml_line1 = str(ws.cell(row=1, column=1).value or "<?xml version=\"1.0\" encoding=\"utf-8\"?>")
    xml_line2 = str(ws.cell(row=2, column=1).value or "")
    wb.close()

    log_func(f"Generation XML Chants (EdgeBands)...")

    eb_txt, eb_count = _export_vba_xml_sheet(xlsm_path, "EdgeBands", log_func=log_func)
    log_func(f"  {eb_count} chants lus")

    # Assembler le fichier
    eb_body = eb_txt.split("\r\n", 2)[2] if "\r\n" in eb_txt else ""
    full_xml = (xml_line1 + "\r\n" + xml_line2 + "\r\n" + eb_body
                + "\r\n</SWOODMat>")
    _validate_xml_document(full_xml)

    # Ecriture du fichier
    if output_dir is None:
        output_dir = os.path.dirname(os.path.abspath(xlsm_path))

    filename, output_path = _write_unique_export(
        output_dir, "Import_Swood_Chants", ".xml", full_xml)

    log_func(f"Fichier cree : {filename}")

    return output_path


# ---------------------------------------------------------------------------
# Interface graphique
# ---------------------------------------------------------------------------

class App:
    """Interface graphique - Theme Destribois 2024"""

    # Palette Destribois (identique a DestriChiffrage)
    PRIMARY = '#2E3544'
    PRIMARY_LIGHT = '#3D4556'
    SECONDARY = '#AE9367'
    SECONDARY_LIGHT = '#C4AB82'
    ACCENT = '#3B7A57'
    ACCENT_LIGHT = '#4A9068'
    BG = '#F7F5F2'
    BG_ALT = '#FFFFFF'
    BG_DARK = '#EDE6DC'
    TEXT = '#2E3544'
    TEXT_LIGHT = '#5A6270'
    TEXT_MUTED = '#8A8F98'
    BORDER = '#DDD8D0'
    DANGER = '#B85450'
    WHITE = '#FFFFFF'

    # Polices Destribois
    FONT_HEADING = ('Abhaya Libre SemiBold', 16)
    FONT_BODY = ('Roboto', 11)
    FONT_BODY_BOLD = ('Roboto Medium', 11)
    FONT_SMALL = ('Roboto', 10)
    FONT_MONO = ('Consolas', 10)

    def __init__(self):
        _load_gui_dependencies()
        self.root = tk.Tk()
        self.root.title(f"Export Optiplanning & SWOOD v{APP_VERSION} - Destribois")
        self.root.geometry("720x660")
        self.root.resizable(True, True)
        self.root.configure(bg=self.BG)

        # --- Header Destribois avec logo ---
        header = tk.Frame(self.root, bg=self.PRIMARY, height=60)
        header.pack(fill="x")
        header.pack_propagate(False)

        header_content = tk.Frame(header, bg=self.PRIMARY)
        header_content.pack(fill="both", expand=True, padx=20)

        # Logo Destribois a gauche (ratio preserve, hauteur adaptee au header)
        self._logo_img = None
        logo_candidates = [
            r"Y:\01_EURL Destribois\10_Communication\01_Charte_graphique\Logo\Logo_Destribois_seul.png",
        ]
        if getattr(sys, 'frozen', False):
            logo_candidates.insert(0, os.path.join(sys._MEIPASS, "Logo_Destribois_seul.png"))
        for logo_path in logo_candidates:
            try:
                pil_img = Image.open(logo_path)
                target_h = 44
                ratio = target_h / pil_img.height
                target_w = int(pil_img.width * ratio)
                pil_img = pil_img.resize((target_w, target_h), Image.LANCZOS)
                self._logo_img = ImageTk.PhotoImage(pil_img)
                tk.Label(header_content, image=self._logo_img,
                         bg=self.PRIMARY).pack(side="left", padx=(0, 12), pady=6)
                break
            except Exception:
                continue

        tk.Label(header_content, text=f"Export Optiplanning & SWOOD v{APP_VERSION}",
                 font=self.FONT_HEADING,
                 fg=self.WHITE, bg=self.PRIMARY).pack(side="left", pady=10)

        # "DESTRIBOIS" a droite - police Abhaya Libre SemiBold, couleur doree #AE9367
        tk.Label(header_content, text="DESTRIBOIS",
                 font=('Abhaya Libre SemiBold', 14),
                 fg=self.SECONDARY, bg=self.PRIMARY).pack(side="right", pady=10)

        # --- Separateur or ---
        tk.Frame(self.root, bg=self.SECONDARY, height=3).pack(fill="x")

        # --- Zone contenu principale ---
        main_frame = tk.Frame(self.root, bg=self.BG, padx=20, pady=12)
        main_frame.pack(fill="both", expand=True)

        # --- Card : Fichier source ---
        src_card = tk.Frame(main_frame, bg=self.BG_ALT, padx=16, pady=12,
                            highlightbackground=self.BORDER, highlightthickness=1)
        src_card.pack(fill="x", pady=(0, 8))

        tk.Label(src_card, text="Fichier source XLSM",
                 font=self.FONT_BODY_BOLD, bg=self.BG_ALT,
                 fg=self.TEXT).pack(anchor="w")

        src_path_frame = tk.Frame(src_card, bg=self.BG_ALT)
        src_path_frame.pack(fill="x", pady=(6, 0))

        self.path_var = tk.StringVar()
        self.path_entry = tk.Entry(src_path_frame, textvariable=self.path_var,
                                   font=self.FONT_SMALL, bg=self.BG_ALT, fg=self.TEXT,
                                   bd=1, relief="solid",
                                   highlightbackground=self.BORDER,
                                   highlightcolor=self.ACCENT,
                                   highlightthickness=1)
        self.path_entry.pack(side="left", fill="x", expand=True)

        btn_browse_src = tk.Button(src_path_frame, text="Parcourir",
                                   command=self.browse_source,
                                   font=self.FONT_SMALL, bg=self.BG_DARK, fg=self.TEXT,
                                   activebackground=self.BORDER,
                                   bd=0, padx=12, pady=4, cursor="hand2", relief="flat")
        btn_browse_src.pack(side="right", padx=(8, 0))

        # --- Card : Dossier destination ---
        dst_card = tk.Frame(main_frame, bg=self.BG_ALT, padx=16, pady=12,
                            highlightbackground=self.BORDER, highlightthickness=1)
        dst_card.pack(fill="x", pady=(0, 8))

        tk.Label(dst_card, text="Dossier de destination",
                 font=self.FONT_BODY_BOLD, bg=self.BG_ALT,
                 fg=self.TEXT).pack(anchor="w")

        dst_path_frame = tk.Frame(dst_card, bg=self.BG_ALT)
        dst_path_frame.pack(fill="x", pady=(6, 0))

        self.output_var = tk.StringVar()
        self.output_entry = tk.Entry(dst_path_frame, textvariable=self.output_var,
                                     font=self.FONT_SMALL, bg=self.BG_ALT, fg=self.TEXT,
                                     bd=1, relief="solid",
                                     highlightbackground=self.BORDER,
                                     highlightcolor=self.ACCENT,
                                     highlightthickness=1)
        self.output_entry.pack(side="left", fill="x", expand=True)

        btn_browse_dst = tk.Button(dst_path_frame, text="Parcourir",
                                   command=self.browse_output,
                                   font=self.FONT_SMALL, bg=self.BG_DARK, fg=self.TEXT,
                                   activebackground=self.BORDER,
                                   bd=0, padx=12, pady=4, cursor="hand2", relief="flat")
        btn_browse_dst.pack(side="right", padx=(8, 0))

        self.default_label = tk.Label(dst_card,
                 text="(vide = meme dossier que le XLSM)",
                 font=('Roboto', 9), bg=self.BG_ALT, fg=self.TEXT_MUTED)
        self.default_label.pack(anchor="w", pady=(2, 0))

        # --- Card : Exports ---
        export_card = tk.Frame(main_frame, bg=self.BG_ALT, padx=16, pady=12,
                               highlightbackground=self.BORDER, highlightthickness=1)
        export_card.pack(fill="x", pady=(0, 8))

        tk.Label(export_card, text="Exports disponibles",
                 font=self.FONT_BODY_BOLD, bg=self.BG_ALT,
                 fg=self.TEXT).pack(anchor="w", pady=(0, 8))

        # Bouton 1 : TXT Optiplanning (primary/accent)
        self.btn_txt = self._create_btn(
            export_card, "Export TXT Optiplanning",
            "8 colonnes tab-delimited - Page Materials",
            self.ACCENT, self.ACCENT_LIGHT, self.do_export_txt)

        # Bouton 2 : XML Plaques Nesting (secondary/or)
        self.btn_nesting = self._create_btn(
            export_card, "Export XML Plaques Nesting",
            "Boards pour SWOOD Nesting - Page Materials",
            self.SECONDARY, self.SECONDARY_LIGHT, self.do_export_nesting)

        # Bouton 3 : XML Materiaux (primary/bleu-gris)
        self.btn_materials = self._create_btn(
            export_card, "Export XML Materiaux SWOOD",
            "Materiaux complets 49 colonnes - Page Materials",
            self.PRIMARY, self.PRIMARY_LIGHT, self.do_export_materials)

        # Bouton 4 : XML Chants (secondary darker)
        self.btn_edgebands = self._create_btn(
            export_card, "Export XML Chants (EdgeBands)",
            "Chants pour SWOOD - Page EdgeBands",
            '#8A7652', '#A08B66', self.do_export_edgebands)

        # --- Card : Journal ---
        log_card = tk.Frame(main_frame, bg=self.BG_ALT, padx=16, pady=12,
                            highlightbackground=self.BORDER, highlightthickness=1)
        log_card.pack(fill="both", expand=True)

        tk.Label(log_card, text="Journal",
                 font=self.FONT_BODY_BOLD, bg=self.BG_ALT,
                 fg=self.TEXT).pack(anchor="w", pady=(0, 6))

        self.log_text = tk.Text(log_card, font=self.FONT_MONO, height=10,
                                bg=self.BG, fg=self.TEXT,
                                relief="solid", borderwidth=1,
                                highlightbackground=self.BORDER,
                                highlightthickness=0,
                                insertbackground=self.TEXT,
                                selectbackground=self.ACCENT,
                                selectforeground=self.WHITE)
        self.log_text.pack(fill="both", expand=True)

        # --- Barre de statut en bas ---
        status_bar = tk.Frame(self.root, bg=self.PRIMARY, height=32)
        status_bar.pack(fill="x", side="bottom")
        status_bar.pack_propagate(False)

        self.status_var = tk.StringVar(value="Pret")
        self.status_label = tk.Label(status_bar, textvariable=self.status_var,
                                     font=self.FONT_SMALL, fg=self.WHITE,
                                     bg=self.PRIMARY, anchor="w", padx=16)
        self.status_label.pack(fill="both", expand=True)

        # --- Init ---
        self._find_default_xlsm()
        self._all_buttons = [self.btn_txt, self.btn_nesting,
                             self.btn_materials, self.btn_edgebands]

        self.log("Pret. Selectionnez un fichier XLSM et choisissez un export.")

    def _create_btn(self, parent, text, subtitle, bg, hover_bg, command):
        """Cree un bouton d'export style Destribois avec sous-titre."""
        frame = tk.Frame(parent, bg=self.BG_ALT)
        frame.pack(fill="x", pady=2)

        btn = tk.Button(frame, text=text, command=command,
                        font=self.FONT_BODY_BOLD, bg=bg, fg=self.WHITE,
                        activebackground=hover_bg, activeforeground=self.WHITE,
                        bd=0, padx=16, pady=8, cursor="hand2", relief="flat",
                        anchor="w")
        btn.pack(side="left", fill="x", expand=True)

        # Sous-titre a droite dans le bouton
        lbl = tk.Label(frame, text=subtitle, font=('Roboto', 9),
                       bg=self.BG_ALT, fg=self.TEXT_MUTED)
        lbl.pack(side="right", padx=(8, 0))

        # Hover effect
        def on_enter(e):
            btn.configure(bg=hover_bg)
        def on_leave(e):
            btn.configure(bg=bg)
        btn.bind('<Enter>', on_enter)
        btn.bind('<Leave>', on_leave)

        return btn

    def _find_default_xlsm(self):
        """Cherche le XLSM dans le meme dossier que l'exe/script."""
        if getattr(sys, 'frozen', False):
            base = os.path.dirname(sys.executable)
        else:
            base = os.path.dirname(os.path.abspath(__file__))

        candidate = os.path.join(base, "Outil_Material_Import.xlsm")
        if os.path.exists(candidate):
            self.path_var.set(candidate)
        else:
            for f in os.listdir(base):
                if f.endswith(".xlsm") and "backup" not in f.lower() and "copie" not in f.lower():
                    self.path_var.set(os.path.join(base, f))
                    break

    def browse_source(self):
        path = filedialog.askopenfilename(
            title="Selectionner le fichier XLSM",
            filetypes=[("Fichiers Excel Macro", "*.xlsm"), ("Tous", "*.*")],
        )
        if path:
            self.path_var.set(path)

    def browse_output(self):
        folder = filedialog.askdirectory(title="Dossier de destination")
        if folder:
            self.output_var.set(folder)

    def log(self, msg):
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")
        self.root.update_idletasks()

    def _set_status(self, msg, color=None):
        """Met a jour la barre de statut en bas de fenetre."""
        if color is None:
            color = self.WHITE
        self.status_var.set(msg)
        self.status_label.configure(fg=color)

    def _get_xlsm_path(self) -> Optional[str]:
        """Valide et retourne le chemin XLSM."""
        xlsm = self.path_var.get().strip()
        if not xlsm:
            self._set_status("Selectionnez un fichier XLSM.", self.SECONDARY)
            return None
        if not os.path.exists(xlsm):
            self._set_status(f"Fichier introuvable : {xlsm}", self.DANGER)
            return None
        return xlsm

    def _get_output_dir(self) -> Optional[str]:
        """Retourne le dossier de destination (ou None = meme que XLSM)."""
        out = self.output_var.get().strip()
        if out and os.path.isdir(out):
            return out
        return None

    def _disable_buttons(self):
        for btn in self._all_buttons:
            btn.config(state="disabled")

    def _enable_buttons(self):
        for btn in self._all_buttons:
            btn.config(state="normal")

    def _run_export(self, export_func, export_name):
        """Wrapper generique pour tous les exports."""
        xlsm = self._get_xlsm_path()
        if not xlsm:
            return

        output_dir = self._get_output_dir()

        self._disable_buttons()
        self._set_status(f"Export en cours : {export_name}...", self.SECONDARY)
        self.log_text.delete("1.0", "end")
        self.log(f"=== {export_name} ===")
        self.log(f"Source : {os.path.basename(xlsm)}")
        if output_dir:
            self.log(f"Destination : {output_dir}")
        else:
            self.log(f"Destination : {os.path.dirname(os.path.abspath(xlsm))}")
        self.log("")

        try:
            result = export_func(xlsm, output_dir=output_dir, log_func=self.log)
            if result:
                self.log("")
                self.log(f"Export termine avec succes !")
                self.log(f"Fichier : {result}")
                self._set_status(
                    f"Export reussi : {os.path.basename(result)}",
                    self.ACCENT)
            else:
                self.log("ERREUR : L'export a echoue.")
                self._set_status("Echec de l'export. Voir le journal.", self.DANGER)
        except PermissionError:
            self.log("ERREUR : Fichier verrouille. Fermez Excel et reessayez.")
            self._set_status(
                "Erreur : fichier verrouille. Fermez Excel et reessayez.",
                self.DANGER)
        except Exception as e:
            self.log(f"ERREUR : {e}")
            import traceback
            self.log(traceback.format_exc())
            self._set_status(f"Erreur : {e}", self.DANGER)
        finally:
            self._enable_buttons()

    def do_export_txt(self):
        self._run_export(export_optiplanning_txt, "Export TXT Optiplanning")

    def do_export_nesting(self):
        self._run_export(export_xml_boards_nesting, "Export XML Plaques Nesting")

    def do_export_materials(self):
        self._run_export(export_xml_materials, "Export XML Materiaux SWOOD")

    def do_export_edgebands(self):
        self._run_export(export_xml_edgebands, "Export XML Chants (EdgeBands)")

    def run(self):
        self.root.mainloop()


# ---------------------------------------------------------------------------
# Point d'entree
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    # Mode ligne de commande si argument
    if len(sys.argv) > 1:
        xlsm = sys.argv[1]
        if not os.path.exists(xlsm):
            print(f"ERREUR : Fichier introuvable : {xlsm}")
            sys.exit(1)

        export_type = sys.argv[2] if len(sys.argv) > 2 else "txt"

        if export_type == "txt":
            result = export_optiplanning_txt(xlsm)
        elif export_type == "nesting":
            result = export_xml_boards_nesting(xlsm)
        elif export_type == "materials":
            result = export_xml_materials(xlsm)
        elif export_type == "edgebands":
            result = export_xml_edgebands(xlsm)
        else:
            print(f"Type d'export inconnu : {export_type}")
            print("Types valides : txt, nesting, materials, edgebands")
            sys.exit(1)

        sys.exit(0 if result else 1)

    # Mode GUI
    app = App()
    app.run()
